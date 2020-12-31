import json
import re
import dateparser
import calendar
import requests
import logging
from nameparser import HumanName
from openpyxl import load_workbook
import argparse
import sqlite3
from inflection import parameterize
from os import path
import time
import googlemaps

# command line arguments
parser = argparse.ArgumentParser(description='Spreadsheet to JSON')
parser.add_argument('-input', type=str, help='spreadsheet in xslx format')
parser.add_argument('-key', type=str, help='google api key')
parser.add_argument('-sheet', type=str, help='worksheet to transform')
args = parser.parse_args()
workbook = load_workbook(filename=args.input, data_only=True)

GOOGLE_API_KEY = args.key
gmaps = googlemaps.Client(key=GOOGLE_API_KEY)

# sheet = workbook.active
sheet = workbook[args.sheet]

# logging config
timestr = time.strftime("%Y%m%d-%H%M%S")
logging.basicConfig(filename='logs/import-' + timestr + '.csv', filemode='a', format='%(message)s')
logging.warning('"id","message"')

do_geocode_birth = False
do_geocode_death = False

# spreadsheet column constants

NAME_SALUTATION = 29
NAME_FIRST = 2
NAME_MIDDLE = 26
NAME_LAST = 1
MAIDEN_NAME = 25
GENDER = 24

BURIAL_PLACE = 12
BURIAL_PLOT = 13
BURIAL_CEMETERY_INDEX = 14

BIRTH_CITY = 8
BIRTH_COUNTY = 9
BIRTH_STATE = 11
# BIRTH_COUNTRY =

OBITUARY_DAY = 3
OBITUARY_MONTH = 4
OBITUARY_YEAR = 5
OBITUARY_NEWSPAPER = 6
OBITUARY_TRANSCRIBED = 28

BIRTH_DAY = 7
BIRTH_MONTH = 10
BIRTH_YEAR = 0

ADMIN_CODE_NUMBER = 15
ADMIN_DATE_RECORD_ADDED = 16
ADMIN_DATE_RECORD_CHANGED = 17
ADMIN_NOTES = 27

DEATH_CITY = 18
DEATH_COUNTY = 19
DEATH_STATE = 22
# DEATH_COUNTRY =

DEATH_MONTH = 21
DEATH_DAY = 20
DEATH_YEAR = 23

TAGS = 30


# =====================================================================================================================
# GOOGLE GEOCODE API LOOKUP
# =====================================================================================================================
def get_google_geocode_results(address_or_zipcode):
    base_url = "https://maps.googleapis.com/maps/api/geocode/json"
    endpoint = f"{base_url}?address={address_or_zipcode}&key={api_key}"
    # see how our endpoint includes our API key? Yes this is yet another reason to restrict the key
    r = requests.get(endpoint)
    if r.status_code not in range(200, 299):
        return None, None
    try:
        '''
        This try block in case any of our inputs are invalid. This is done instead
        of actually writing out handlers for all kinds of responses.
        '''
        results = r.json()['results'][0]
    except:
        pass
    return results


def get_geocode_dict():
    d = {}
    keys = [
        'place_geocode_input',
        'geo_lat', 'geo_lng',
        'geo_formatted_address',
        'geo_street_number',
        'geo_street_name_long',
        'geo_street_name_short',
        'geo_neighborhood',
        'geo_country_long',
        'geo_country_short',
        'geo_city',
        'geo_state_long',
        'geo_state_short',
        'geo_county',
        'geo_zip',
        'geo_other',
        'google_place_id'
    ]
    # initialize empty dictionary
    for i in keys:
        d[i] = ''
    return d

# =====================================================================================================================
# GEOCODE PLACE
# =====================================================================================================================
def geocode_place(id, type, place):

    d = get_geocode_dict()
    # return empty dictionary if no place is specified
    if place == '':
        return d
    else:
        d['place_geocode_input'] = place

    geocode_filename = 'json/places/' + parameterize(place.lower().strip(), separator="_") + '.json'
    if place != '' and path.exists(geocode_filename) is False:

        # geocode_results = get_google_geocode_results(place)
        gmaps_result = gmaps.geocode(place)
        if len(gmaps_result) != 0:
            geocode_results = gmaps_result[0]
        else:
            geocode_results = None

        if geocode_results is not None:
            d['google_place_id'] = geocode_results['place_id']
            d['place_geocode_input'] = place
            d['geo_lat'] = geocode_results['geometry']['location']['lat']
            d['geo_lng'] = geocode_results['geometry']['location']['lng']
            d['geo_formatted_address'] = geocode_results['formatted_address']
            for component in geocode_results['address_components']:
                for geo_type in component['types']:
                    if geo_type == 'street_number':
                        d['geo_street_number'] = component['short_name']
                    if geo_type == 'route':
                        d['geo_street_name_long'] = component['long_name']
                        d['geo_street_name_short'] = component['short_name']
                    if geo_type == 'neighborhood':
                        d['geo_neighborhood'] = component['long_name']
                    if geo_type == 'country':
                        d['geo_country_short'] = component['short_name']
                        d['geo_country_long'] = component['long_name']
                    if geo_type == 'sublocality':
                        d['geo_city'] = component['long_name']
                    if geo_type == 'locality':
                        d['geo_city'] = component['long_name']
                    if geo_type == 'administrative_area_level_1':
                        d['geo_state_long'] = component['long_name']
                        d['geo_state_short'] = component['short_name']
                    if geo_type == 'administrative_area_level_2':
                        d['geo_county'] = component['long_name']
                    if geo_type == 'postal_code':
                        d['geo_zip'] = component['long_name']
                    if geo_type == 'establishment':
                        d['geo_other'] = component['long_name']
                    if geo_type == 'natural_feature':
                        d['geo_other'] = component['long_name']
            try:
                f = open(geocode_filename, 'w')
                json.dump(d, f, indent=4, sort_keys=True)
                f.close()
                return d
            except Exception as e:
                logging.fatal(getattr(e, 'message', repr(e)))
        else:
            logging.warning('"' + str(int(id)) + '","' + "Unable geocode " + type + " place: " + place + '"')
            # serialize empty geo place
            f = open(geocode_filename, 'w')
            json.dump(d, f, indent=4, sort_keys=True)
            f.close()
            return d
    else:
        # return serialized geocodes
        if path.exists(geocode_filename):
            with open(geocode_filename) as json_file:
                place_json = json.load(json_file)
                if place_json["google_place_id"] == "":
                    logging.warning('"' + str(int(id)) + '","' + "Unable geocode " + type + " place: " + place + '"')
                return place_json
        else:
            return d

# =====================================================================================================================
# IS FLOAT?
# =====================================================================================================================
def isfloat(value):
  try:
    float(value)
    return True
  except ValueError:
    return False

# =====================================================================================================================
# MAIN
# =====================================================================================================================
entries = []
current_row = 0

# Using the values_only because you want to return the cells' values
for row in sheet.iter_rows(min_row=2, values_only=True):

    # --- ROW ID
    current_row = current_row + 1
    id = current_row

    # --- TAGS
    tags = []
    if len(row) > TAGS:
        if row[TAGS] is not None:
            tags = re.findall(r'"(.*?)"', row[TAGS])

    # --- NAME
    name_salutation = ''
    name_first = ''
    gender = 'Unknown'
    name_middle = ''
    name_last = ''
    name_maiden = ''
    name_full = ''
    if row[NAME_SALUTATION] is not None:
        name_salutation = str(row[NAME_SALUTATION]).strip().capitalize()
    if row[NAME_FIRST] is not None:
        name_first = str(row[NAME_FIRST]).strip().capitalize()
    if row[NAME_MIDDLE] is not None:
        name_middle = str(row[NAME_MIDDLE]).strip().capitalize()
    if row[NAME_LAST] is not None:
        name_last = str(row[NAME_LAST]).strip().capitalize()
    if row[MAIDEN_NAME] is not None:
        name_maiden = str(row[MAIDEN_NAME]).strip().capitalize()
    name = HumanName(name_salutation + " " + name_first + " " + name_middle + " " + name_last)
    name_full = name.full_name
    if row[GENDER] is not None:
        gender = str(row[GENDER]).strip().upper()

    #--- BURIAL LOCATION
    burial_place = ''
    burial_plot = ''
    burial_cemetery_index = ''
    if row[BURIAL_PLACE] is not None:
        burial_place = row[BURIAL_PLACE].strip()
    if row[BURIAL_PLOT] is not None:
        burial_plot = row[BURIAL_PLOT].strip()
    if row[BURIAL_CEMETERY_INDEX] is not None:
        burial_cemetery_index = str(row[BURIAL_CEMETERY_INDEX]).strip()

    # --- PLACE OF BIRTH
    birth_city = ''
    birth_county = ''
    birth_state = ''
    birth_country = ''
    birth_place_full = ''
    birth_geo_country_short = ''
    birth_geo_country_long = ''
    birth_geo_city = ''
    birth_geo_state_long = ''
    birth_geo_state_short = ''
    birth_geo_county = ''
    birth_geo_zip = ''
    birth_geo_lat = None
    birth_geo_lng = None
    birth_geo_formatted_address = ''

    # BIRTH CITY
    if row[BIRTH_CITY] is not None:
        birth_city = str(row[BIRTH_CITY]).strip().title()

    # BIRTH COUNTY
    if row[BIRTH_COUNTY] is not None:
        birth_county = str(row[BIRTH_COUNTY]).strip().title()

    # BIRTH STATE
    if row[BIRTH_STATE] is not None:
        birth_state = str(row[BIRTH_STATE]).strip().upper()

#         # BIRTH COUNTRY
#         if row[BIRTH_COUNTRY] is not None:
#             birth_country = str(row[BIRTH_COUNTRY])

    # BIRTH PLACE FULL - concatenate all of the available fields
    birth_place_full = (birth_city + ", " + birth_state)

    # GEOCODE: BIRTH PLACE
    if do_geocode_birth:
        birth_place_geocoded = geocode_place(id, 'birth', birth_place_full)
    else:
        birth_place_geocoded = get_geocode_dict()

    # --- PLACE OF DEATH
    death_location = ''
    death_street = ''
    death_city = ''
    death_county = ''
    death_state = ''
    death_country = ''
    death_place_full = ''
    death_geo_country_short = ''
    death_geo_country_long = ''
    death_geo_city = ''
    death_geo_state_long = ''
    death_geo_state_short = ''
    death_geo_county = ''
    death_geo_zip = ''
    death_geo_lat = None
    death_geo_lng = None
    death_geo_formatted_address = ''
    if row[DEATH_CITY] is not None:
        death_city = str(row[DEATH_CITY]).strip().title()
    if row[DEATH_COUNTY] is not None:
        death_county = str(row[DEATH_COUNTY]).strip().title()
    if row[DEATH_STATE] is not None:
        death_state = str(row[DEATH_STATE]).strip().upper()
#         if row[DEATH_COUNTRY] is not None:
#             death_country = str(row[DEATH_COUNTRY]).strip()
    death_place_full = (death_city + ", " + death_state)

    # GEOCODE: DEATH PLACE
    if do_geocode_death:
        death_place_geocoded = geocode_place(id, 'death', death_place_full)
    else:
        death_place_geocoded = get_geocode_dict()

    # --- DEATH DATE
    death_date_display = ''
    death_date_iso = None
    death_year = ''
    death_year_int = None
    death_circa = 'N'
    # death month
    if row[DEATH_MONTH] is not None and row[DEATH_MONTH] != '':
        death_month = str(row[DEATH_MONTH]).strip()
        m = re.search('0?(\d+)', death_month)
        if m == None:
            logging.warning('"' + str(int(id)) + '","' + "Unable to parse death month: " + death_month + '"')
        else:
            death_month_int = int(m.group(1))
            if death_month_int >= 1 and death_month_int <= 12:
                death_month_str = calendar.month_name[death_month_int]
                death_date_display += death_month_str + " "
            else:
                logging.warning('"' + str(int(id)) + '","' + "Unable to parse death month: " + m.group(1) + '"')
    # death day
    if row[DEATH_DAY] is not None and row[DEATH_DAY] != '':
        # correct capital letter O used instead of zero
        death_day = str(row[DEATH_DAY]).strip()
        death_day = death_day.replace("O","0").replace(".0", "")
        death_date_display += death_day + ", "
    if "@" in death_date_display:
        death_circa = 'Y'
    # year
    if row[DEATH_YEAR] is not None and row[DEATH_YEAR] != '':
        death_year = str(row[DEATH_YEAR]).strip()
        if "@" in death_year:
            death_circa = 'Y'
        m = re.search('(\d\d\d\d)', death_year)
        if m == None:
            logging.warning('"' + str(int(id)) + '","' + "Unable to parse death year: " + death_year + '"')
        else:
            death_year_int = int(m.group(1))
            death_date_display += m.group(1)
            if death_date_display != '':
                dt = dateparser.parse(death_date_display)
                if dt is not None:
                    death_date_iso = dt.strftime("%Y-%m-%d")
                else:
                    logging.warning('"' + str(int(id)) + '","' + "Unable to parse death date: " + death_date_display + '"')

    # --- BIRTH DATE
    birth_date_display = ''
    birth_date_iso = None
    birth_year = ''
    birth_month = ''
    birth_day = ''
    birth_circa = 'N'
    birth_year_int = None
    # month
    if row[BIRTH_MONTH] is not None and row[BIRTH_MONTH] != '':
        birth_month = str(row[BIRTH_MONTH]).strip()
        m = re.search('0?(\d+)', birth_month)
        if m == None:
            logging.warning('"' + str(int(id)) + '","' + "Unable to parse birth month: " + birth_month + '"')
        else:
            birth_month_int = int(m.group(1))
            if birth_month_int >= 1 and birth_month_int <= 12:
                birth_month_str = calendar.month_name[birth_month_int]
                birth_date_display += birth_month_str + " "
            else:
                logging.warning('"' + str(int(id)) + '","' + "Unable to parse birth month: " + m.group(1) + '"')
    # day
    if row[BIRTH_DAY] is not None and row[BIRTH_DAY] != '':
        birth_day = str(row[BIRTH_DAY]).strip()
        birth_date_display += death_day + ", "
    if "@" in birth_date_display:
        birth_circa = 'Y'
    # year
    if row[BIRTH_YEAR] is not None and row[BIRTH_YEAR] != '':
        birth_year = str(row[BIRTH_YEAR]).strip()
        if "@" in birth_year:
            birth_circa = 'Y'
        m = re.search('(\d\d\d\d)', birth_year)
        if m == None:
            logging.warning('"' + str(int(id)) + '","' + "Unable to parse birth year: " + birth_year + '"')
        else:
            birth_year_int = int(m.group(1))
            birth_date_display += m.group(1)
            if birth_date_display != '':
                dt = dateparser.parse(birth_date_display)
                if dt is not None:
                    birth_date_iso = dt.strftime("%Y-%m-%d")
                else:
                    logging.warning('"' + str(int(id)) + '","' + "Unable to parse birth date: " + birth_date_display + '"')

    # --- OBITUARY
    obituary_date_display = ''
    obituary_date_iso = None
    obituary_year = ''
    obituary_year_int = None
    obituary_month = ''
    obituary_day = ''
    obituary_newspaper = ''
    obituary_transcribed = ''
    obituary_circa = 'N'
    # month
    if row[OBITUARY_MONTH] is not None and row[OBITUARY_MONTH] != '':
        obituary_month = str(row[OBITUARY_MONTH]).strip()
        m = re.search('0?(\d+)', obituary_month)
        if m == None:
            logging.warning('"' + str(int(id)) + '","' + "Unable to parse obituary month: " + obituary_month + '"')
        else:
            obituary_month_int = int(m.group(1))
            if obituary_month_int >= 1 and obituary_month_int <= 12:
                obituary_month_str = calendar.month_name[obituary_month_int]
                obituary_date_display += obituary_month_str + " "
            else:
                logging.warning('"' + str(int(id)) + '","' + "Unable to parse birth month: " + m.group(1) + '"')

    # day
    if row[OBITUARY_DAY] is not None and row[OBITUARY_DAY] != '':
        obituary_day = str(row[OBITUARY_DAY]).strip()
        obituary_date_display += obituary_day + ", "
    if "@" in obituary_date_display:
        obituary_circa = 'Y'
    # year
    if row[OBITUARY_YEAR] is not None and row[OBITUARY_YEAR] != '':
        obituary_year = str(row[OBITUARY_YEAR]).strip()
        if "@" in death_year:
            obituary_circa = 'Y'
        m = re.search('(\d\d\d\d)', obituary_year)
        if m == None:
            logging.warning('"' + str(int(id)) + '","' + "Unable to parse obituary year: " + obituary_year + '"')
        else:
            obituary_year_int = int(m.group(1))
            obituary_date_display += m.group(1)
            if obituary_date_display != '':
                dt = dateparser.parse(obituary_date_display)
                if dt is not None:
                    obituary_date_iso = dt.strftime("%Y-%m-%d")
                else:
                    logging.warning('"' + str(int(id)) + '","' + "Unable to parse obituary date: " + obituary_date_display + '"')
    if row[OBITUARY_NEWSPAPER] is not None:
        obituary_newspaper = str(row[OBITUARY_NEWSPAPER]).strip()
    if row[OBITUARY_TRANSCRIBED] is not None:
        obituary_transcribed = str(row[OBITUARY_TRANSCRIBED]).strip()

    # --- ADMIN
    admin_notes = ''
    admin_date_added = ''
    admin_date_changed = ''
    admin_code_number = ''
    if row[ADMIN_DATE_RECORD_ADDED] is not None:
        admin_date_added = str(row[ADMIN_DATE_RECORD_ADDED]).strip()
    if row[ADMIN_DATE_RECORD_CHANGED] is not None:
        admin_date_changed = str(row[ADMIN_DATE_RECORD_CHANGED]).strip()
    if row[ADMIN_NOTES] is not None:
        admin_notes = str(row[ADMIN_NOTES]).strip()
    if row[ADMIN_CODE_NUMBER] is not None:
        admin_code_number = str(row[ADMIN_CODE_NUMBER]).strip()

    entry = {
        "id": id,
        "name_salutation": name_salutation,
        "name_first": name_first,
        "name_middle": name_middle,
        "name_last": name_last,
        "name_full": name_full,
        "name_maiden": name_maiden,
        "gender": gender,
        "burial_place": burial_place,
        "burial_plot": burial_plot,
        "burial_cemetery_index": burial_cemetery_index,
        "birth_city": birth_city,
        "birth_county": birth_county,
        "birth_state": birth_state,
        "birth_place_full": birth_place_full,
        "birth_geo_location": {"lat": birth_place_geocoded['geo_lat'],"lon": birth_place_geocoded['geo_lng']},
        "birth_geo_city": birth_place_geocoded['geo_city'],
        "birth_geo_county": birth_place_geocoded['geo_county'],
        "birth_geo_state_short": birth_place_geocoded['geo_state_short'],
        "birth_geo_state_long": birth_place_geocoded['geo_state_long'],
        "birth_geo_country_long": birth_place_geocoded['geo_country_long'],
        "birth_geo_country_short": birth_place_geocoded['geo_country_short'],
        "birth_geo_zip": birth_place_geocoded['geo_zip'],
        "birth_geo_place_id": birth_place_geocoded['google_place_id'],
        "birth_geo_formatted_address": birth_place_geocoded['geo_formatted_address'],
        "death_city": death_city,
        "death_county": death_county,
        "death_state": death_state,
        "death_country": death_country,
        "death_place_full": death_place_full,
        "death_geo_location": {"lat": death_place_geocoded['geo_lat'],"lon": death_place_geocoded['geo_lng']},
        "death_geo_city": death_place_geocoded['geo_city'],
        "death_geo_county": death_place_geocoded['geo_county'],
        "death_geo_state_short": death_place_geocoded['geo_state_short'],
        "death_geo_state_long": death_place_geocoded['geo_state_long'],
        "death_geo_country_long": death_place_geocoded['geo_country_long'],
        "death_geo_country_short": death_place_geocoded['geo_country_short'],
        "death_geo_zip": death_place_geocoded['geo_zip'],
        "death_geo_place_id": death_place_geocoded['google_place_id'],
        "death_geo_formatted_address": death_place_geocoded['geo_formatted_address'],
        "death_date_display" : death_date_display,
        "death_date_iso": death_date_iso,
        "death_date_year": death_year_int,
        "death_date_circa": death_circa,
        "birth_date_display" : birth_date_display,
        "birth_date_iso": birth_date_iso,
        "birth_date_year": birth_year_int,
        "birth_date_circa": birth_circa,
        "obituary_date_display" : obituary_date_display,
        "obituary_date_iso": obituary_date_iso,
        "obituary_date_year": obituary_year_int,
        "obituary_date_circa": obituary_circa,
        "obituary_newspaper": obituary_newspaper,
        "obituary_transcribed": obituary_transcribed,
        "admin_notes": admin_notes,
        "admin_date_added": admin_date_added,
        "admin_date_changed": admin_date_changed,
        "admin_code_number": admin_code_number,
        "tags": tags
    }
    # make sure you can json serialize, otherwise dump error and interment
    try:
        json_temp = json.dumps(entry)
        entries.append(entry)
    except Exception as e:
        logging.fatal(getattr(e, 'message', repr(e)))
        logging.fatal(entry)
        # exit()

print(json.dumps(entries))